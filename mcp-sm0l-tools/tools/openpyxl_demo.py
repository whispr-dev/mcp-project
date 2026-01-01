"""openpyxl create workbook demo"""

def main():
    try:
        import tempfile, pathlib
    from openpyxl import Workbook
    except Exception as e:
        print("Missing dependency or import failed:", e)
        return
    wb = Workbook(); ws = wb.active
    ws["A1"] = "hello"; ws["B1"] = 123
    p = pathlib.Path(tempfile.gettempdir())/"sm0l_openpyxl_demo.xlsx"
    wb.save(p)
    print("wrote", p)

if __name__ == "__main__":
    main()
